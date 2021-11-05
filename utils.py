# utils.py is just a function created to help format data in Excel
# Not part of the workflow (and not used much in these later iterations of the code)

import openpyxl as xl  # openpyxl is an excel manipulation library
###  Help:   https://openpyxl.readthedocs.io/en/stable/tutorial.html
from openpyxl.utils import get_column_interval
import re
import pandas as pd
import pdb
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import (
    ScatterChart,
    Reference,
    Series,
)
from openpyxl.styles import NamedStyle, Font
from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder
from openpyxl.utils import get_column_letter

def row_color_fill( ws, row, color):
    for cell in ws[f"{row}:{row}"]:
        # cell.style = 'Headline 4'
        cell.fill = xl.styles.PatternFill(start_color=color, end_color=color,
                                          fill_type="solid")
def row_bold( ws, row, bold):
    for cell in ws[f"{row}:{row}"]:
        # cell.style = 'Headline 4'
        cell.font = Font(bold = bold)

def autowidth_all_cols(wsheet):
    dim_holder = DimensionHolder(worksheet=wsheet)

    for col in range(wsheet.min_column, wsheet.max_column + 1):
        dim_holder[get_column_letter(col)] = ColumnDimension(wsheet, min=col, max=col, width=20)

    wsheet.column_dimensions = dim_holder

def load_workbook_range(range_string, ws):
    col_start, col_end = re.findall("[A-Z]+", range_string)

    data_rows = []
    for row in ws[range_string]:
        data_rows.append([cell.value for cell in row])

    return pd.DataFrame(data_rows, columns=get_column_interval(col_start, col_end))

def hasNoNumbers(row):
    inputString = row['Ticker']
    return (not any(char.isdigit() for char in inputString))

def place_dataframe_at(df, wsheet, first_row, first_col, col_formats, col_styles):
    idx_y = 0
    idx_r = first_row
    for r in dataframe_to_rows(df, index=False, header=False):
        idx_x = 0
        idx_c = first_col

        for col in r:
            wsheet.cell(row = idx_r, column = idx_c).value = col

            if idx_y > -1:
                if col_formats[idx_x] != '':
                    wsheet.cell(row=idx_r, column=idx_c).number_format = col_formats[idx_x]
                if col_styles[idx_x] != '':
                    wsheet.cell(row=idx_r, column=idx_c).style = col_styles[idx_x]
            idx_c = idx_c + 1
            idx_x = idx_x + 1
        idx_y = idx_y + 1
        idx_r = idx_r + 1

def write_row_list_at( data_list, wsheet, row, first_col, styles, overwrite = True):
    idx_c = first_col
    idx_r = row
    for item in data_list:
        cell = wsheet.cell(row=idx_r, column=idx_c)
        if overwrite == True:
            cell.value = item
        if 'bold' in styles:
            cell.font = Font(bold = styles['bold'] )
        if 'fill' in styles:
            cell.fill = xl.styles.PatternFill(start_color=styles['fill'], end_color=styles['fill'],
                                              fill_type="solid")
        if 'number_format' in styles:
            cell.number_format = styles['number_format']
        idx_c = idx_c + 1

def paste_dataframe_at(df, wsheet, first_row, first_col):
    idx_y = 0
    idx_r = first_row
    for r in dataframe_to_rows(df, index=False, header=False):
        idx_x = 0
        idx_c = first_col

        for col in r:
            wsheet.cell(row = idx_r, column = idx_c).value = col


            idx_c = idx_c + 1
            idx_x = idx_x + 1
        idx_y = idx_y + 1
        idx_r = idx_r + 1