# program reads individual files created by get_data_from_bloomberg.py and aggregates them
# compiled dataframe is stored in this csv
dataframe_csv_filename ="./data/compiled_dataframe.csv"
CREATE_NEW_CSV = True
# Write here the list of files to aggregate - to work, each file must first be individually processed by get_data_from_bloomberg.py
copypaste_files_list  = [
    "./data/copypastePaloAlto.xlsx",
    "./data/copypasteDAFNA",
    "./data/copypasteBakerBros",
    # "copypasteBVF.xlsx",
    # "copypastePerceptiveNew",
    # "copypasteRA",
    # "copypasteRTW",
    # "copypasteRedmile",
    # "copypasteEcoR1",
    # "copypasteOrbimed",
    # "copypasteAvoro",
    # "copypasteCormorant",
    # "copypasteBoxer",
    #
    ]

import openpyxl as xl  # openpyxl is an excel manipulation library
import pandas as pd
import numpy as np
import utils
import pickle
import os

if CREATE_NEW_CSV:
    try:
        os.remove(dataframe_csv_filename)
    except:
        pass


try:
    df = pd.read_csv(dataframe_csv_filename) #try reading an existing csv
except:
    df = pd.DataFrame(columns = ['Fund Name', 'Name', 'Date', 'Ticker', 'Market Cap', 'Market Val Wgt', 'Market Val',
                       'Position', 'PX Close', 'PX Last', 'Prev Date', 'PX Last Prev Date', 'PX Ratio', 'Next Date',
                       'PX Last Next Date', 'PX Ratio Next']) #...if it doesn't exist, create the dataframe df

# run through various raw and pickle files one at a time
for copypaste_file in copypaste_files_list:

    print(f"Processing {copypaste_file}")
    read_from_file = "./data/"+ copypaste_file.split('.')[1].split('/')[-1] + "_raw.xlsx"
    get_ranges_from_file =  "./data/"+ copypaste_file.split('.')[1].split('/')[-1]  + "_ranges.pickle"

    ranges = pickle.load(open(get_ranges_from_file,'rb'))

    # open an existing xlsx
    wb_in = xl.load_workbook(filename = read_from_file, read_only = True)
    wb_in_data = wb_in['Data']

    for idx,r in enumerate(ranges):
        int_start = r['report_row_start']
        int_stop = r['report_row_stop']
        print('New Range')

        df1 = utils.load_workbook_range(f'A{int_start}:P{int_stop}', wb_in_data)
        df1.columns = ['Name', 'tobedeleted', 'Date', 'Ticker', 'Market Cap', 'Market Val Wgt', 'Market Val',
                       'Position', 'PX Close', 'PX Last', 'Prev Date', 'PX Last Prev Date', 'PX Ratio', 'Next Date',
                       'PX Last Next Date', 'PX Ratio Next'] #define a preliminary dataframe, using existing columns
        df1 = df1.drop(columns=['tobedeleted'])

        df_fundname = utils.load_workbook_range(f'A{int_start-1}:A{int_start-1}', wb_in_data)    #for each raw copypaste file, the fundname is (-1,-1) relative to the data
        fundname = df_fundname['A'].values[0]   #

        df1['Fund Name'] = fundname #append 'Fund Name' to the preliminary df1

        df = df.append(df1) #add this column to the processed dataframe df, that we are going to use-- this will automatically add the data to the already existing 'Fund Name' column in df

    wb_in.close()

df.to_csv(dataframe_csv_filename, index = False)




