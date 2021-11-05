
#Variable defining where initial data comes from
get_data_for_copypaste_file = "./data/copypasteBakerBros.xlsx"

#Variable defining where final data goes to
put_data_in_file = "./data/"+ get_data_for_copypaste_file.split('.')[1].split('/')[-1]+ "_raw.xlsx"
put_ranges_in_file = "./data/" + get_data_for_copypaste_file.split('.')[1].split('/')[-1] + "_ranges.pickle" # .pickle is a file type that stores information for Python

import openpyxl as xl  # openpyxl is an excel manipulation library
###  Help:   https://openpyxl.readthedocs.io/en/stable/tutorial.html
from openpyxl.utils import get_column_interval
import re
import pandas as pd
import bl_getdata as av_bdh
import time
import pdb
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import (
    ScatterChart,
    Reference,
    Series,
)
from openpyxl.styles import NamedStyle, Font
from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder
from openpyxl.utils import get_column_letter
import numpy as np
import utils
import pickle



# open "Sheet1"  from the copypaste xlsx, print an error if it doesn't work
try:
    wb_in = xl.load_workbook(filename = get_data_for_copypaste_file, read_only = True) # define wb_in by the ws loaded in as ... copypaste
except:
    print(f"Unable to open file {get_data_for_copypaste_file}") # if it doesn't work, print an error message

try:
    ws_in = wb_in['Sheet1'] #Now, define wb_in as 'Sheet1' from workbook previously defined as wb_in
except:
    print(f'No sheet titled "Sheet1" in {get_data_for_copypaste_file}')


#create a new workbook
wb_out = xl.Workbook() #wb_out is the output wb
ws_data = wb_out['Sheet'] # change the name of the default sheet
ws_data.title = "Data" # change the name of the selected tab to "Data"


#Scans the copypaste sheet line by line, and defines a set of ranges in the Excel rows that correspond to each period data set
idx = 0 #default values
ranges = []
in_range = True
range_start = 0

for row in ws_in.rows:
    idx = idx+1
    if (not any(cell.value for cell in row)): #scanning across, iterates through cell values in the row
        if idx == 1:
            in_range = False
        if not in_range:
            in_range = True
            range_start = idx
        else:
            in_range = False
            range_stop = idx
            ranges.append( {
                'range_start': range_start+1,
                'range_stop': range_stop-1
            })
        # print(f"Row {idx} is empty")
    if (idx == ws_in.max_row) and in_range: # .max_row refers to the end of a chunk
        range_stop = idx
        ranges.append({
            'range_start': range_start+1,
            'range_stop': range_stop
        })

print(ranges)


# get all dates; works by finding the cell that contains the date, and removing all other scrap information
all_dates = []
for r in ranges:
    ticker_date = ws_in[f"A{r['range_start'] + 1}"].value
    d = ticker_date.split()[-1].split('/')
    dd = '20' + d[2] + d[0] + d[1]
    all_dates.append(dd)

print(all_dates)
dates_prev = all_dates[1::]
dates_next = all_dates[0:-1]
dates_prev.append(None)
dates_next.insert(0,None)


# get neccesary data from bloomberg, and write everything in Excel
idx_out = 0
idx_sorting = 0
df_changes = pd.DataFrame(columns = ['Date','P/L% Change'])
for r, this_date, prev_date, next_date in zip(ranges, all_dates, dates_prev, dates_next):
    idx = 0
    r_idx = 0
    idx_sorting = idx_sorting + 2
    idx_out = idx_out + 2 # skip two rows
    ticker_cells = ws_in[f"J{r['range_start'] + 6}:J{r['range_stop']}"]
    # ticker_date = active_sheet[f"A{r['range_start'] + 1}"].value
    tickers = [x[0].value + ' EQUITY' for x in ticker_cells]
    # d = ticker_date.split()[-1].split('/')
    # dd = '20' + d[2] + d[0] + d[1]
    print('POINT1')
    ticker_data = av_bdh.getData(tickers, this_date)
    print('POINT2')
    if prev_date is not None:
        ticker_data_prev = av_bdh.getData(tickers, prev_date)
    else:
        ticker_data_prev = ticker_data
    if next_date is not None:
        ticker_data_next = av_bdh.getData(tickers, next_date)
    else:
        ticker_data_next = ticker_data
    print(ticker_data)
    for row in ws_in.rows:
        idx = idx+1
        if (idx >=r['range_start']) and (idx <= r['range_stop']):
            r_idx = r_idx + 1
            if r_idx == 2:
                r['position_date'] = row[0].value.split()[-1]
                idx_out = idx_out + 1
                ws_data[f"A{idx_out}"] = r['position_date']
                ws_data[f"A{idx_out}"].font = xl.styles.Font(bold=True)
                idx_out = idx_out + 1
                ws_data[f"A{idx_out}"] = "Name"
                ws_data[f"B{idx_out}"] = "#"
                ws_data[f"C{idx_out}"] = "Date"
                ws_data[f"D{idx_out}"] = "Ticker Equity"
                ws_data[f"E{idx_out}"] = "Market Cap"
                ws_data[f"F{idx_out}"] = "% Wgt"
                ws_data[f"G{idx_out}"] = "Market Val"
                ws_data[f"H{idx_out}"] = "Pos"
                ws_data[f"I{idx_out}"] = "PX Close"
                ws_data[f"J{idx_out}"] = "PX Last"
                ws_data[f"K{idx_out}"] = "Prev Date"
                ws_data[f"L{idx_out}"] = "PX Last Prev Date"
                ws_data[f"M{idx_out}"] = "PX Ratio Prev"
                ws_data[f"N{idx_out}"] = "Next Date"
                ws_data[f"O{idx_out}"] = "PX Last Next Date"
                ws_data[f"P{idx_out}"] = "PX Ratio Next"

                utils.row_bold(ws_data, idx_out, True)

            if r_idx == 6: # fund name and summary
                r['fund_name'] = row[2].value #Refers to column C in Excel (copypaste)
                r['fund_number'] = row[3].value
                r['fund_total_percentage'] = row[4].value
                r['fund_market_val'] = row[5].value
                idx_out = idx_out + 2
                ws_data[f"A{idx_out}"] = r['fund_name']
                ws_data[f"B{idx_out}"] = r['fund_number']
                ws_data[f"F{idx_out}"] = r['fund_total_percentage']
                ws_data[f"G{idx_out}"] = r['fund_market_val']

                utils.row_bold(ws_data, idx_out, True)
            if r_idx >= 7: # data starts here
                idx_out = idx_out + 1
                r['Ticker Equity'] = row[9].value + ' EQUITY'
                ws_data[f"C{idx_out}"] = r['position_date']
                ws_data[f"A{idx_out}"] = row[2].value
                ws_data[f"D{idx_out}"] = r['Ticker Equity']
                # try:
                #     d = r['position_date'].split('/')
                #     dd = '20'+d[2] + d[1] + d[0]
                #     mkt_cap_data = av_bdh.getData(r['Ticker Equity'], dd)[0]['cur_mkt_val']
                #     time.sleep(0.5)
                # except:
                #     mkt_cap_data = None
                mkt_cap_data = list(filter(lambda ticker: ticker['name'] == r['Ticker Equity'], ticker_data))[0]['cur_mkt_val']
                px_last = list(filter(lambda ticker: ticker['name'] == r['Ticker Equity'], ticker_data))[0]['px_last']
                px_last_prev_date = list(filter(lambda ticker: ticker['name'] == r['Ticker Equity'], ticker_data_prev))[0]['px_last']
                px_last_next_date = list(filter(lambda ticker: ticker['name'] == r['Ticker Equity'], ticker_data_next))[0]['px_last']
                # print(r['Ticker Equity'], mkt_cap_data, dd)
                ws_data[f"E{idx_out}"] = mkt_cap_data
                ws_data[f"J{idx_out}"] = px_last
                ws_data[f"L{idx_out}"] = px_last_prev_date
                ws_data[f"K{idx_out}"] = prev_date
                if (px_last_prev_date is not None) and (px_last is not None):
                    px_ratio_prev = float(px_last)/float(px_last_prev_date)
                else:
                    px_ratio_prev = None
                ws_data[f"M{idx_out}"] = px_ratio_prev

                if (px_last_next_date is not None) and (px_last is not None):
                    px_ratio_next = float(px_last_next_date)/float(px_last)
                else:
                    px_ratio_next = None
                ws_data[f"P{idx_out}"] = px_ratio_next
                ws_data[f"O{idx_out}"] = px_last_next_date
                ws_data[f"N{idx_out}"] = next_date
                # ws_data[f"E{idx_out}"] = row[5].value
                ws_data[f"F{idx_out}"] = row[4].value
                ws_data[f"G{idx_out}"] = row[5].value
                ws_data[f"H{idx_out}"] = row[6].value
                ws_data[f"I{idx_out}"] = row[7].value
                if r_idx == 7:
                    r['report_row_start'] = idx_out
    r['report_row_stop'] = idx_out

    int_start = int(r['report_row_start'])
    int_stop = int(r['report_row_stop'])

wb_out.save(filename=put_data_in_file)  # wil overwrite if a file of same name exists
wb_out.close()
wb_in.close()

fn = open(put_ranges_in_file, 'wb')
pickle.dump(ranges, fn)
fn.close()
